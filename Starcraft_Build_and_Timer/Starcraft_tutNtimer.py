from PyQt5.QtWidgets import *
from PyQt5 import uic
import sys
import openpyxl
import os
import time

from PyQt5.QtCore import Qt

ui_file = "./starcraft_ui.ui"
check_find_b = True
build_name_input = False
build_output_index = 0
nowb_idx = 3
start_check = True
time_stamp = []
begin = 0
last_output = 0


def check_xlsfile():
    if not(os.path.exists('./BuildNLog.xlsx')):
        bnl_open = openpyxl.Workbook()
        bnl_open.save("./BuildNLog.xlsx")
        bnl = openpyxl.load_workbook("./BuildNLog.xlsx")
        bnl_sheet = bnl.active
        bnl_sheet['A1'] = 2
        bnl_sheet['B1'] = 2
        bnl_sheet['C1'] = "1열은 수정하지 말아주세요"
        bnl.save("./BuildNLog.xlsx")

class MainDialog(QDialog):
    def __init__(self):
        QDialog.__init__(self,None)
        uic.loadUi(ui_file,self)
        self.setWindowTitle("Starcraft Build tut & timer")
        self.initUI()
        self.main_view()
        check_xlsfile()

        #load_build
        bnl = openpyxl.load_workbook("./BuildNLog.xlsx")
        bnl_sheet = bnl.active
        if bnl_sheet["B1"].value + 1 > bnl_sheet["A1"].value:
            bnl_sheet["A1"] = bnl_sheet["B1"].value
        check_upper = True
        for i in range(2, (bnl_sheet['A1'].value) + 1):
            if bnl_sheet['A{}'.format(i)].value != None:
                self.pick_b.addItem(bnl_sheet['A{}'.format(i)].value)
            elif bnl_sheet['A{}'.format(i)].value == None and check_upper == True:
                bnl_sheet["B1"] = i
                check_upper = False
        bnl.save("./BuildNLog.xlsx")

        self.find_b.clicked.connect(self.find_b_click)
        self.input_b.clicked.connect(self.input_b_click)
        self.tab_key.clicked.connect(self.build_tab)
        self.alt_key.clicked.connect(self.build_alt)

    def initUI(self):
        self.pick_b.addItem("새로운 빌드 만들기")
        self.tab_key.setShortcut("Tab")
        self.alt_key.setShortcut("Q")

    def main_view(self):
        self.textBrowser.clear()
        self.textBrowser.append("< 빌드 생성 >\n")
        self.textBrowser.append("1. \"새로운 빌드 만들기\" 선택")
        self.textBrowser.append("2. 빈칸에 빌드 제목 입력 후 입력 버튼 클릭")
        self.textBrowser.append("3. 빌드 내용 (ex 인구수 9 오버로드 ) 입력\n")
        self.textBrowser.append("\n< 생성된 빌드 선택 >\n")
        self.textBrowser.append("1. 박스에서 원하는 빌드 제목 선택")
        self.textBrowser.append("2. 선택 버튼 클릭\n")
        self.textBrowser.append("\n< 생성된 빌드 삭제 >\n")
        self.textBrowser.append("1. 삭제하고 싶은 빌드 선택 후 선택 버튼 클릭")
        self.textBrowser.append("2. 입력 창에 \"삭제\" 입력 후 입력 버튼 클릭")
        self.textBrowser.append("\n< 선택된 빌드 종료 후 로그 확인 >\n")
        self.textBrowser.append("\'새로운 빌드 만들기\" 선택 후 선택 버튼 두번 클릭")
        self.textBrowser.append("\n< 빌드 타이머 및 로그 기능 >\n")
        self.textBrowser.append("게임 시작시 바로 tab키 눌러 타이머 작동 필수")
        self.textBrowser.append("tab 키 --> 빌드 순서 넘기기")
        self.textBrowser.append("Q 키 --> 빌드 순서 되돌리기\n")

    def show_time(self):
        global time_stamp
        bnl = openpyxl.load_workbook("./BuildNLog.xlsx")
        bnl_sheet = bnl.active
        self.textBrowser.append(" < 빌드 로그 및 걸린 시간 확인 >\n")
        for i in range(0,len(time_stamp)):
            if bnl_sheet.cell(last_output, i+3).value != None:
                self.textBrowser.append("{}번 - ".format(i+1) + str(bnl_sheet.cell(last_output,i+3).value) + " : " +str(time_stamp[i]) + " 초")
        time_stamp = []

    def input_b_click(self):
        global build_name_input
        if check_find_b == False:
            buildxl = openpyxl.load_workbook("./BuildNLog.xlsx")
            b_sheet = buildxl.active
            build_idx = b_sheet['B1'].value
            if build_name_input == True and self.pick_b.currentIndex() == 0:
                textinput = self.edit_b.toPlainText()
                b_sheet['A{}'.format(build_idx)] = textinput
                b_sheet['B{}'.format(build_idx)] = 3
                self.pick_b.addItem(textinput)
                self.label.setText("제목 입력 완료 !")
                build_name_input = False
            elif self.pick_b.currentIndex() == 0 and build_name_input == False:
                b_sheet.cell(build_idx, b_sheet["B{}".format(build_idx)].value,self.edit_b.toPlainText())
                b_sheet["B{}".format(build_idx)] = b_sheet["B{}".format(build_idx)].value + 1
                self.label.setText("빌드 입력 완료 !")
            buildxl.save("./BuildNLog.xlsx")
        elif self.pick_b.currentIndex() != 0 and check_find_b == True and self.edit_b.toPlainText() == "삭제":
            buildxl = openpyxl.load_workbook("./BuildNLog.xlsx")
            bnl_sheet = buildxl.active
            for i in range(2,(bnl_sheet["A1"].value)+1):
                if self.pick_b.currentText() == bnl_sheet["A{}".format(i)].value:
                    bnl_sheet.delete_rows(i)
                    self.label.setText("삭제 완료 !")
                    if self.pick_b.currentIndex() != 0:
                        self.pick_b.removeItem(self.pick_b.currentIndex() )
                    break
            buildxl.save("./BuildNLog.xlsx")

    def find_b_click(self):
        global check_find_b
        global build_name_input
        global build_output_index
        global nowb_idx
        global start_check
        global last_output
        if self.pick_b.currentIndex() == 0 and check_find_b == True:
            self.label.setText("빌드 제목 입력 후 빌드 순서를 입력\n저장&종료는 선택버튼 클릭 (이후 다른빌드 선택가능)\n빌드 저장을 위해 바로 다른 빌드 선택 불가")
            check_find_b = False
            build_name_input = True
            last_output = build_output_index
            build_output_index = 0
            nowb_idx = 3
        elif self.pick_b.currentIndex() == 0 and check_find_b == False:
            check_find_b = True
            build_name_input = False
            self.label.clear()
            bnl = openpyxl.load_workbook("./BuildNLog.xlsx")
            bnl_sheet = bnl.active
            bnl_sheet["B1"] = bnl_sheet["B1"].value + 1
            while True:
                if bnl_sheet['A{}'.format(bnl_sheet["B1"].value)].value != None:
                    bnl_sheet["B1"] = bnl_sheet["B1"].value + 1
                else:
                    break
            if bnl_sheet["B1"].value + 1 > bnl_sheet["A1"].value:
                bnl_sheet["A1"] = bnl_sheet["B1"].value
            for i in range(2, (bnl_sheet['A1'].value) + 1):
                if bnl_sheet['A{}'.format(i)].value == None:
                    bnl_sheet["B1"] = i
                    break
            bnl.save("./BuildNLog.xlsx")
            main_dialog.main_view()
            main_dialog.show_time()
        elif self.pick_b.currentIndex() != 0 and check_find_b == True:
            self.textBrowser.clear()
            self.label.clear()
            start_check = True
            nowb_idx = 3
            bnl = openpyxl.load_workbook("./BuildNLog.xlsx")
            bnl_sheet = bnl.active
            for i in range(2,(bnl_sheet["A1"].value)+1):
                if self.pick_b.currentText() == bnl_sheet["A{}".format(i)].value:
                    build_output_index = i
                    break
            self.textBrowser.append("빌드 제목 : < " + bnl_sheet.cell(build_output_index,1).value + " >")
            self.textBrowser.append("시작시 tab키 누르고 시작")
            self.textBrowser.append("해당 단계 완료시 tab키 누르기")
            self.textBrowser.append("이후 로그에서 해당 단계 완료시점 확인 가능\n")
            for t in range(3,bnl_sheet["B{}".format(build_output_index)].value):
                self.textBrowser.append("{} : ".format(t-2) + str(bnl_sheet.cell(build_output_index,t).value))

    def build_tab(self):
        global nowb_idx
        global build_output_index
        global start_check
        global begin
        global time_stamp
        if build_output_index != 0:
            if start_check == True:
                time_stamp = []
                begin = time.time()
                start_check = False
            else:
                stamp = round(time.time() - begin, 3)
                time_stamp.append(stamp)
        bnl = openpyxl.load_workbook("./BuildNLog.xlsx")
        bnl_sheet = bnl.active
        if build_output_index == 0:
            boi = 1
        else:
            boi = build_output_index
        if int(bnl_sheet["B{}".format(boi)].value) > nowb_idx and build_output_index != 0:
            nowb_idx += 1
        if build_output_index != 0:
            if bnl_sheet.cell(build_output_index, nowb_idx - 1).value == None:
                first = "빌드 내용이 없습니다."
            else:
                first = str(bnl_sheet.cell(build_output_index, nowb_idx - 1).value)
            if bnl_sheet.cell(build_output_index, nowb_idx).value == None:
                second = "빌드 내용이 없습니다."
            else:
                second = str(bnl_sheet.cell(build_output_index, nowb_idx).value)
            if bnl_sheet.cell(build_output_index, nowb_idx + 1).value == None:
                third = "빌드 내용이 없습니다."
            else:
                third = str(bnl_sheet.cell(build_output_index, nowb_idx + 1).value)
            self.label.setText(first + "<-현재 순서\n" + second + "\n" + third)

    def build_alt(self):
        global nowb_idx
        global build_output_index
        if len(time_stamp) > 0 and build_output_index != 0:
            time_stamp.pop()
        bnl = openpyxl.load_workbook("./BuildNLog.xlsx")
        bnl_sheet = bnl.active
        if 4 < nowb_idx:
            nowb_idx -= 1
        if build_output_index != 0:
            if bnl_sheet.cell(build_output_index, nowb_idx - 1).value == None:
                first = "빌드 내용이 없습니다."
            else:
                first = str(bnl_sheet.cell(build_output_index, nowb_idx - 1).value)
            if bnl_sheet.cell(build_output_index, nowb_idx).value == None:
                second = "빌드 내용이 없습니다."
            else:
                second = str(bnl_sheet.cell(build_output_index, nowb_idx).value)
            if bnl_sheet.cell(build_output_index, nowb_idx + 1).value == None:
                third = "빌드 내용이 없습니다."
            else:
                third = str(bnl_sheet.cell(build_output_index, nowb_idx + 1).value)
            self.label.setText(first + "<-현재 순서\n" + second + "\n" + third)

def my_exception_hook(exctype, value, traceback):
    # Print the error and traceback
    print(exctype, value, traceback)
    # Call the normal Exception hook after
    sys._excepthook(exctype, value, traceback)
    # sys.exit(1)

# Back up the reference to the exceptionhook
sys._excepthook = sys.excepthook

# Set the exception hook to our wrapping function
sys.excepthook = my_exception_hook


QApplication.setStyle("fusion")
app = QApplication(sys.argv)
main_dialog = MainDialog()
main_dialog.show()
sys.exit(app.exec_())